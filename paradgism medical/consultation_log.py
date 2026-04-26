"""
consultation_log.py -- Excel-based consultation log.

Each call to record() appends a new sheet to consultations.xlsx.
Sheet name = "<PatientName> <DD-Mon-YYYY>" (31-char Excel limit enforced).
Each sheet contains a structured table: vitals, symptoms, diagnosis, candidates.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

dataDir = Path(__file__).parent / "patient_data"
dataDir.mkdir(exist_ok=True)

# ── Style constants ────────────────────────────────────────────────────────────
hdrFill    = PatternFill("solid", fgColor="1E3A5F")
hdrFont    = Font(bold=True, color="FFFFFF", size=11)
secFill    = PatternFill("solid", fgColor="DBEAFE")
secFont    = Font(bold=True, color="1E3A5F", size=10)
boldFont   = Font(bold=True, size=10)
normalFont = Font(size=10)
centerAlign = Alignment(horizontal="center", vertical="center", wrap_text=True)
leftAlign   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def thinBorder():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def sheetName(patient_name: str, dt: datetime) -> str:
    date_part = dt.strftime("%d-%b-%Y")
    name_part = patient_name.strip()[:18]
    full = f"{name_part} {date_part}"
    return full[:31]


def setCell(ws, row: int, col: int, value, font=None, fill=None,
              alignment=None, border=None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    if font:      cell.font      = font
    if fill:      cell.fill      = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border


def sectionHeader(ws, row: int, text: str, span: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    setCell(ws, row, 1, text, font=secFont, fill=secFill, alignment=leftAlign)


def writeKv(ws, row: int, key: str, value) -> None:
    setCell(ws, row, 1, key,   font=boldFont,   alignment=leftAlign, border=thinBorder())
    setCell(ws, row, 2, value, font=normalFont, alignment=leftAlign, border=thinBorder())


def record(
    patient_name:   str,
    session_id:     str,
    age:            Optional[int],
    vitals_summary: str,
    news2Score:     int,
    news2Risk:      str,
    confirmed:      List[str],
    denied:         List[str],
    skipped:        List[str],
    diagnosis:      str,
    confidence:     float,
    severity:       int,
    severity_label: str,
    candidates:     List[Dict],
    questions_asked: int,
    durationSeconds: Optional[int],
    gender:         str = "Rather not say",
    visit_dt:       Optional[datetime] = None,
    file_path:      Optional[Path] = None,
) -> Path:
    """Write one consultation Excel file and return its path."""

    dt       = visit_dt or datetime.now()
    out_path = file_path or (dataDir / f"consultation_{session_id}.xlsx")

    # Each consultation gets its own fresh workbook
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    sheet_name = sheetName(patient_name, dt)
    # Avoid duplicate sheet names by appending a counter
    base = sheet_name
    counter = 2
    while sheet_name in wb.sheetnames:
        sheet_name = f"{base[:28]} {counter}"
        counter += 1

    ws = wb.create_sheet(sheet_name)

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    row = 1

    # ── Title banner ───────────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    setCell(ws, row, 1, "MEDICAL CONSULTATION RECORD",
              font=Font(bold=True, color="FFFFFF", size=13),
              fill=PatternFill("solid", fgColor="1E3A5F"),
              alignment=centerAlign)
    ws.row_dimensions[row].height = 26
    row += 1

    # ── Patient info ───────────────────────────────────────────────────────────
    sectionHeader(ws, row, "PATIENT INFORMATION", 4); row += 1
    writeKv(ws, row, "Patient Name",   patient_name);       row += 1
    writeKv(ws, row, "Age",            age if age else "Not provided"); row += 1
    writeKv(ws, row, "Gender",         gender);              row += 1
    writeKv(ws, row, "Session ID",     session_id);          row += 1
    writeKv(ws, row, "Date & Time",    dt.strftime("%Y-%m-%d  %H:%M")); row += 1
    writeKv(ws, row, "Duration",       f"{durationSeconds}s" if durationSeconds else "N/A"); row += 1
    writeKv(ws, row, "Questions Asked", questions_asked);    row += 1
    row += 1

    # ── Vitals / NEWS2 ─────────────────────────────────────────────────────────
    sectionHeader(ws, row, "VITAL SIGNS  (NEWS2)", 4); row += 1
    writeKv(ws, row, "Readings",   vitals_summary);          row += 1
    writeKv(ws, row, "NEWS2 Score", news2Score);            row += 1
    writeKv(ws, row, "NEWS2 Risk",  news2Risk);             row += 1
    row += 1

    # ── Diagnosis ──────────────────────────────────────────────────────────────
    sectionHeader(ws, row, "DIAGNOSIS", 4); row += 1
    writeKv(ws, row, "Primary Diagnosis",
        diagnosis.replace("_", " ").upper() if diagnosis != "undetermined" else "UNDETERMINED")
    row += 1
    writeKv(ws, row, "Confidence",    f"{confidence:.1%}");  row += 1
    writeKv(ws, row, "Severity Score", severity);            row += 1
    writeKv(ws, row, "Severity Level", severity_label);      row += 1
    row += 1

    # ── Differential diagnoses table ───────────────────────────────────────────
    if candidates:
        sectionHeader(ws, row, "DIFFERENTIAL DIAGNOSES", 4); row += 1

        # Table headers
        for col_i, hdr in enumerate(["Rank", "Disease", "Confidence", "Notes"], start=1):
            setCell(ws, row, col_i, hdr,
                      font=hdrFont, fill=hdrFill,
                      alignment=centerAlign, border=thinBorder())
        tbl_start = row
        row += 1

        for rank, cand in enumerate(candidates[:5], start=1):
            d = cand.get("disease", "unknown")
            c = cand.get("confidence", 0.0)
            note = "Primary" if rank == 1 else ""
            setCell(ws, row, 1, rank,                      font=normalFont, alignment=centerAlign, border=thinBorder())
            setCell(ws, row, 2, d.replace("_", " ").title(), font=normalFont, alignment=leftAlign,   border=thinBorder())
            setCell(ws, row, 3, f"{c:.1%}",                font=normalFont, alignment=centerAlign, border=thinBorder())
            setCell(ws, row, 4, note,                      font=boldFont,   alignment=centerAlign, border=thinBorder())
            row += 1

        # Excel Table
        tbl_ref = f"A{tbl_start}:D{row - 1}"
        tbl = Table(displayName=f"Diff_{session_id}", ref=tbl_ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        ws.add_table(tbl)
        row += 1

    # ── Symptoms ───────────────────────────────────────────────────────────────
    sectionHeader(ws, row, "SYMPTOMS", 4); row += 1

    # Confirmed
    if confirmed:
        setCell(ws, row, 1, f"Confirmed  ({len(confirmed)})",
                  font=Font(bold=True, color="166534", size=10),
                  fill=PatternFill("solid", fgColor="DCFCE7"),
                  alignment=leftAlign, border=thinBorder())
        sym_text = ",  ".join(s.replace("_", " ").title() for s in sorted(confirmed))
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        setCell(ws, row, 2, sym_text, font=normalFont, alignment=leftAlign, border=thinBorder())
        row += 1

    # Denied
    if denied:
        setCell(ws, row, 1, f"Denied  ({len(denied)})",
                  font=Font(bold=True, color="991B1B", size=10),
                  fill=PatternFill("solid", fgColor="FEE2E2"),
                  alignment=leftAlign, border=thinBorder())
        sym_text = ",  ".join(s.replace("_", " ").title() for s in sorted(denied))
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        setCell(ws, row, 2, sym_text, font=normalFont, alignment=leftAlign, border=thinBorder())
        row += 1

    # Skipped
    if skipped:
        setCell(ws, row, 1, f"Unknown / Skipped  ({len(skipped)})",
                  font=Font(bold=True, color="374151", size=10),
                  fill=PatternFill("solid", fgColor="F3F4F6"),
                  alignment=leftAlign, border=thinBorder())
        sym_text = ",  ".join(s.replace("_", " ").title() for s in sorted(skipped))
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        setCell(ws, row, 2, sym_text, font=normalFont, alignment=leftAlign, border=thinBorder())
        row += 1

    row += 1

    # ── Disclaimer ─────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    setCell(ws, row, 1,
              "DISCLAIMER: Educational tool only. Always consult a licensed medical professional.",
              font=Font(italic=True, color="EF4444", size=9),
              alignment=centerAlign)

    wb.save(out_path)
    return out_path
